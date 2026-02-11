from django.contrib import admin
from django.urls import path
from django.template.response import TemplateResponse
from datetime import timezone

from django.http import HttpResponse
from django.utils.timezone import now
from .models import Category, Product, Enquiry, Contact
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.html import format_html


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'is_available', 'created_at')
    list_filter = ('category', 'is_available')
    search_fields = ('name',)
    prepopulated_fields = {'slug': ('name',)}




@admin.register(Enquiry)
class EnquiryAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "phone",
        "source",
        "status",
        "product",
        "created_at",
    )

    list_filter = (
        "source",
        "status",
        "created_at",
    )

    search_fields = (
        "name",
        "phone",
        "email",
        "product__name",
    )

    readonly_fields = ("created_at",)

    date_hierarchy = "created_at"

    actions = ["export_to_excel"]
    
    
    def status_badge(self, obj):
        colors = {
            "new" : "blue",
            "contacted" : "orange",
            "quoted" : "purple",
            "negotiation" : "teal",
            "converted" : "green",
            "closed" : "gray",
            "lost" : "red",
        }
        color = colors.get(obj.status, "black")
        return format_html(
            '<span style="color:white;background:{};padding:4px 8px;border-radius:6px;">{}</span>',
            color,
            obj.get_status_display()
        )
    
    status_badge.short_description = "Status"


    def priority_badge(self,obj):
        colors = {
            "low" : "gray",
            "medium" : "orange",
            "high" : "red",
        }
        color = colors.get(obj.priority, "black")
        return format_html(
            '<span style="color:white;background:{};padding:4px 8px;border-radius:6px;">{}</span>',
            color,
            obj.get_priority_display()
        )
        
    priority_badge.short_description = "Priority"
    
    
    def followup_status(self,obj):
        if obj.follow_up_date:
            if obj.follow_up_date < timezone.now().date():
                return format_html('<span style="color:red;font-weight:bold;">Overdue</span>')
            return format_html('<span style="color:green;">Upcoming</span>')
        return "-"
    
    followup_status.short_description = "Follow-Up"
    
    def export_to_excel(self, request, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Enquiries"

        columns = [
            "ID",
            "Name",
            "Phone",
            "Email",
            "Product",
            "Quantity",
            "Source",
            "Status",
            "Priority",
            "Assigned To",
            "Estimated Value",
            "Follow-Up Date"
            "Created At",
        ]

        # Header row
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        # Data rows
        for row_num, enquiry in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1).value = enquiry.id
            worksheet.cell(row=row_num, column=2).value = enquiry.name
            worksheet.cell(row=row_num, column=3).value = enquiry.phone
            worksheet.cell(row=row_num, column=4).value = enquiry.email
            worksheet.cell(row=row_num, column=5).value = (
                enquiry.product.name if enquiry.product else "-"
            )
            worksheet.cell(row=row_num, column=6).value = enquiry.quantity
            worksheet.cell(row=row_num, column=7).value = enquiry.source
            worksheet.cell(row=row_num, column=8).value = enquiry.get_status_display()
            worksheet.cell(row=row_num, column=9).value = enquiry.get_priority_display()
            worksheet.cell(row=row_num, column=10).value = enquiry.assigned_to
            worksheet.cell(row=row_num, column=11).value = (
                float(enquiry.estimated_value) if enquiry.estimated_value else "-"
            )
            worksheet.cell(row=row_num, column=12).value = (
                enquiry.get_follow_up_date.strftime("%Y-%m-%d")
                if enquiry.get_follow_up_date else "-"
            )
            
            
            worksheet.cell(row=row_num, column=9).value = enquiry.created_at.strftime("%Y-%m-%d %H:%M")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=enquiries.xlsx"

        workbook.save(response)
        return response

    export_to_excel.short_description = "📥 Export selected enquiries to Excel"
    
@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):
    list_display = ("name", "phone" , "email", "created_at")
    search_fields = ("name","phone","email")
    ordering = ("-created_at",)
    readonly_fields = ("created_at",)
    
    


admin.site.site_header = "Pashupathinath Marketing Admin"
admin.site.site_title = "PM Admin"
admin.site.index_title = "Dashboard"

def admin_dashboard(request):
    today = now().date()

    context = dict(
        admin.site.each_context(request),

        total_enquiries=Enquiry.objects.count(),
        today_enquiries=Enquiry.objects.filter(created_at__date=today).count(),
        total_products=Product.objects.count(),
        total_categories=Category.objects.count(),
        total_contacts=Contact.objects.count(),

        recent_enquiries=Enquiry.objects.order_by("-created_at")[:10],
    )

    return TemplateResponse(request, "admin/dashboard.html", context)

original_get_urls = admin.site.get_urls

def custom_get_urls():
    urls = original_get_urls()
    custom_urls = [
        path(
            "dashboard/",
            admin.site.admin_view(admin_dashboard),
            name="admin_dashboard"
        ),
    ]
    return custom_urls + urls

admin.site.get_urls = custom_get_urls